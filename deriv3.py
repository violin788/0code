#ssl how to install 1.1.1 vs 1.1.0
#pip install urllib3==1.26.6

#,patoolib
"""
inp =str(input("input = "))
#refresh progress on screw..not new line..print(a,end='\r')

#write txt
new = open(fil,"w")
new.write(tex)
new.close()

tex = rea4([che,'txt'])

cwd = os.getcwd()
lis = os.listdir(fol)
"""
#old import inspect
import json
import time,xlrd,os,sys,csv,pyautogui,time,datetime,xlwt
#import py7zr
import xlsxwriter,subprocess,shutil,math
#import requests
import pandas as pd
#for/to install win32api use instead, pywin32, pip install pywin32, not pip install win32api
import tkinter as tk

def gem(inp):
	#ema = gem([])
	ema = []
	ema.append("media788788@gmail.com")
	#ema.append("desk788788@gmail.com")
	ema.append("violin78@protonmail.com")
	ema.append("cello34@protonmail.com")
	ema.append("violin78@mail.com")
	ema.append("usa2@email.com")
	return ema


def dri_(inp):	
	#dri = dri_([])
	#acc = acc_([])
	dri = os.getcwd()
	dri = dri[0:dri.find("\\")]	
	return dri

def acc_(inp):	
	#dri = dri_([])
	#acc = acc_([])
	loc = os.getcwd()
	loc2 = loc.replace("\\Users\\","")
	acc2 = loc2[loc2.find(":")+1:loc2.find("\\")]
	#dri = dri[0:dri.find("\\")]	
	print(acc2)
	return(acc2)
	
def acc(inp):	
	#acc2 = acc([])
	loc = os.getcwd()
	loc2 = loc.replace("\\Users\\","")
	acc2 = loc2[loc2.find(":")+1:loc2.find("\\")]
	#dri = dri[0:dri.find("\\")]	
	print(acc2)
	return(acc2)



def cuf(inp):
	#fol = cuf([])

	#fol2 = cuf([fol])
	#fol = inp[0]
	
	cwd = os.getcwd()
	cwd2 = rev([cwd])
	cwd3 = cwd2.find("\\")
	cwd4 = cwd2[0:cwd3]
	cwd5 = rev([cwd4])
	print(cwd2)
	print(cwd5)
	return cwd5




def rev(inp):
	#bac = rev([tex])
	tex = inp[0]
	tex2 = ""
	if type(tex)==str:
		ray = []
		for a in range(0,len(tex)):
			val = tex[len(tex)-1-a]
			if val=="\n":
				val = "---\n"
			tex2 = tex2+val
	if type(tex)==list:
		tex2 = tex[::-1]
	return tex2

def rev2(inp):
	#bac = rev2([tex])
	tex = inp[0]
	bac = ""
	for a,val in enumerate(tex):
		bac = bac+tex[len(tex)-1-a]
	return bac

def sen(inp):
	#sen([fil])
	from send2trash import send2trash
	fil = inp[0]
	send2trash(fil)

def inc(inp):
	#inc(["alp"])
	nam = inp[0]
	fil = nam+".txt"
	cwd = os.getcwd()
	#fea =cwd+"\\ear" 
	lis = os.listdir(cwd)
	mat = []
	mat.append(["alp","alpha vantage"])
	mat.append(["yfi","yfinance"])

	from datetime import date
	tod = str(date.today())
	print(tod)
	#fil = "alp.txt"
	tex = rea([fil])
	print("")
	print(tex)
	tex2 = rev([tex])
	print(tex2)
	ide = "---"
	dat = tex2[0:tex2.find(ide)]
	dat2 = rev([dat])
	print(dat2)
	if dat2!=tod:
		rig = tod
		tex3 = tex.replace(dat2,tod)
		fin = tex.find("\n")
		#tex3[0:fin] = "0"
		tex4 = "0"+tex3[fin:len(tex3)]
		print(tex4)
		#fil = "alp2.txt"
		wri2([fil,tex4])
		print("file out = "+fil)
	if dat2==tod:
		print ("same day")	
	
	if fil not in lis:
		tex = ""
		for a,val in enumerate(mat):
			mat2 = val[0]
			if nam==mat2:
				app = val[1]
				out = "0\n"+tex+"\n"+app+"\n"#add date too..
				break
		new = open(fil,"w")
		new.write(out)
		new.close()
	if fil in lis:
		loa = rea([fil])
		fin = loa.find("\n")
		lin = loa[0:fin]
		val = int(lin)
		val2 = val+1
		val3 = str(val2)
		out = val3+loa[fin:len(loa)]
		new = open(fil,"w")
		new.write(out)
		new.close()
		ven = ""
		for a,val in enumerate(mat):
			mat2 = val[0]
			if nam==mat2:
				ven2 = val[1]
				ven = ven2
				print("hits on "+ven+" today = "+val3)
				break

def siz(inp):
	#siz([fil])
	fil = inp[0]
	meg = os.path.getsize(fil)
	meg2 = float(meg)/10000
	meg3 = int(meg2)
	meg4 = float(meg3/100)
	meg5 = "file = "+fil
	meg6 = meg5+"\nsize = "+str(meg4)+"MB"
	print(meg6)
	"""
	siz = int(siz/1000)
	siz = str(float(siz)/1000)+"MB"
	print(siz)
	"""


def sff(inp):
	#sff([])
	sw("firefox.lnk",3)

def sch(inp):
	#sch([3])
	wai = inp[0]
	sw("chrome.lnk",wai)

def goo(inp):
	#sch([3])
	wai = inp[0]
	sw("chrome.lnk",wai)


def ptt(inp):
	#ptt(["tem","py"])
	nam = inp[0]
	typ = inp[1]
	out = nam
	tex = rea4([nam,typ])
	wrt2([tex,out,"txt"])
	tex = rea4([out,"txt"])
	return tex

def gnf(inp):
	#ptt(["tem2","py","tem3","py"])
	in1 = inp[0]
	ex1 = inp[1]
	out1 = inp[2]
	ex2 = inp[3]
	tex = rea4([in1,ex1])
	wrt2([tex,out1,"txt"])
	#tex = rea4([out,"py"])
	return tex


def wrt2(inp):
	#wrt2([tex,nam,ext])
	tex = inp[0]
	nam = inp[1]
	ext = inp[2]
	che = "."+ext
	if che not in nam:
		nam = nam+"."+ext
	new = open(nam,"w")
	new.write(tex)
	new.close()

def wrt(inp):
	#wrt([tex,"name",".txt"])
	tex = inp[0]
	nam = inp[1]
	typ = inp[2]
	fil=nam
	ide="."+typ
	if ide not in nam:
		fil = nam+ide
	fil=fil.replace("..",".")
	print(fil)
	#sye()
	new = open(fil,"w")
	new.write(tex)
	new.close()

def rea5(inp):
	#text = rea5(["aul2","py"])
	nam = inp[0]
	typ = inp[1]
	che = "."+typ
	if che not in nam:
		nam = nam+"."+typ
	tex = open(nam, "r")
	tex2 = tex.read()
	return tex2
	#print(f.read())

def wri2(inp):
	#wri2([fil,dat])
	#wri2([fil,ray])
	#wri2([output_file,earnings_dates])
	fil = inp[0]
	out = inp[1]
	#print("wh")
	#print(out)
	#sye()
	typ = str(type(out))
	#print (typ)
	bac = ""
	tex = [".txt",".py",".html"]
	#out = inp[1]
	for a,val in enumerate(tex):
		if val in fil:
			#sav = fil.replace(val,"2"+val)
			new = open(fil,"w")
			new.write(out)
			new.close()
			break
	if ".xls" in fil:
		to_write = inp[1]
		import xlwt
		from xlwt import Workbook
		new_xls = Workbook()
		add_worksheet = new_xls.add_sheet('Sheet1')
		#writing to worksheet
		for a,val in enumerate(to_write):
			for b,valb in enumerate(val):
				valb = str(valb)
				add_worksheet.write(a,b,valb)
		#column width adjustment
		widths = []
		max_rows = 0
		for a,val in enumerate(to_write):
			check = len(val)
			if check>max_rows:
				max_rows=check
		for a in range(0,max_rows):
			widths.append(0)
		for a,val in enumerate(to_write):
			for b,valb in enumerate(val):
				check = len(str(valb))
				if check>widths[b]:
					widths[b] = check
		#print(widths)
		for a,val in enumerate(widths):
			specific_column = add_worksheet.col(a)
			#column_width.width = (256*check)*2
			specific_column.width = (280*val)
		new_xls.save(fil)
	if ".csv" in fil: 
		"""
		out2 = []
		for a,val in enumerate(out):
			if type(val)==str:
				out2.append([val])
		out = out2
		"""
		if "DataFrame" in typ:
			out.to_csv(fil)  
		if "DataFrame" not in typ:
			#ray = inp[1]
			try:
				csv_file = open(fil, 'wb')
				csv_file.write(out)
				csv_file.close()
			except:
				print("writing .csv failed")
				print("type of data was = "+str(type(fil)))
				#override =str(input("click to continue anyway?"))
	

				

def rea(inp):
	#tex = rea([fil])
	#rea("fil.xlsx")
	fil = inp[0]
	bac = ""
	tex = [".txt",".py",".html"]
	for a,val in enumerate(tex):
		if val in fil:
			loa = open(fil, "r")
			#print(loa)
			loa2 = loa.read()
			bac = loa2
	#print(bac)
	if ".xls" in fil:
		import xlrd
		workbook = xlrd.open_workbook(fil)
		sheet1 = workbook.sheet_by_index(0)	
		bac = []
		for rowNumber in range(sheet1.nrows):
		    row = sheet1.row_values(rowNumber)
		    #print(row)
		    bac.append(row)
	if ".csv" in fil:
		bac = []
		#make dat.csv in the ear folder, not in the cod folder
		with open(fil, newline='') as csvfile:
			#make dat.csv in the ear folder, not in the cod folder
		    reader = csv.reader(csvfile)
		    for row in reader:
		    	bac.append(row)
	return bac


def rea4(inp):
	#text = rea4(["bas3","py"])
	nam = inp[0]
	typ = inp[1]
	fil = nam
	if "."+typ not in nam:
		fil =nam+"."+typ
	#print (fil)
	#print ("fil = "+fil)
	tex = open(fil, "r")
	bac = tex.read()
	return bac

def rec(inp):
	#rec(["mas"])
	nam = inp[0]
	if ".csv" not in nam: 
		nam=nam+".csv"
	mas = []
	#make dat.csv in the ear folder, not in the cod folder
	with open(nam, newline='') as csvfile:
		#make dat.csv in the ear folder, not in the cod folder
	    reader = csv.reader(csvfile)
	    for row in reader:
	    	mas.append(row)
	return mas

def csr(inp):
	#csr(["mas"])
	nam = inp[0]
	if ".csv" not in nam: 
		nam=nam+".csv"
	mas = []
	#make dat.csv in the ear folder, not in the cod folder
	with open(nam, newline='') as csvfile:
		#make dat.csv in the ear folder, not in the cod folder
	    reader = csv.reader(csvfile)
	    for row in reader:
	    	mas.append(row)
	return mas

def csv_reader(inp):
	#price_history = csv_reader([load_file])
	file_to_be_loaded = inp[0]
	if ".csv" not in file_to_be_loaded: 
		file_to_be_loaded=file_to_be_loaded+".csv"
	data_to_be_loaded = []
	#make dat.csv in the ear folder, not in the cod folder
	with open(file_to_be_loaded, newline='') as csvfile:
		#make dat.csv in the ear folder, not in the cod folder
	    reader = csv.reader(csvfile)
	    for row in reader:
	    	data_to_be_loaded.append(row)
	return data_to_be_loaded


def csw(inp):
	#csw([ray,"mas"])
	con = inp[0]
	nam = inp[1]
	if ".csv" not in nam: 
		nam=nam+".csv"
	csv_file = open(nam, 'wb')
	#change so that written into ear folder,not cod folder
	csv_file.write(con)
	csv_file.close()

def wrc(inp):
	#wrc([ray,"mas"])
	con = inp[0]
	nam = inp[1]
	if ".csv" not in nam: 
		nam=nam+".csv"
	csv_file = open(nam, 'wb')
	#change so that written into ear folder,not cod folder
	csv_file.write(con)
	csv_file.close()	


def cen(inp):
	#cen([34,835])
	num1 = inp[0]
	num2 = inp[1]
	div = float(num1)/float(num2)
	div = int(div*10000)
	div = str(float(div/100))+"%"
	return div



def pgd(inp):
	num = inp[0]
	wait = inp[1]
	#for a in range(0,clicks):
	hod3(["ctrl","pagedown",num-1,wait])


def cf_ee(text):
	hod3(["ctrl","f",1,0])
	wri(text,1)
	but(["esc","enter"],0,0)	


def cf_ete(text):
	#cf_ete("log in"):
	hod3(["ctrl","f",1,0])
	wri(text,1)
	but(["esc","tab","enter"],0,0)

def cf_ete2(inp):
	#cf_ete2(["google",2])

	tex = inp[0]
	wai = inp[1]
	#cf_ete("log in"):
	hod3(["ctrl","f",1,0])
	wri(tex,1)
	but(["esc","tab","enter"],0,wai)


def cf_et(text):
	hod3(["ctrl","f",1,0])
	wri(text,1)
	but(["esc","tab"],0,0)

def cf_et2(inp):
	#cf_et2(["email"])
	tex = inp[0]
	hod3(["ctrl","f",1,0])
	wri(tex,1)
	but(["esc","tab"],0,0)




def htb(first_hold,second_hold,button_click,clicks,wait):
	#htb("ctrl","shift",10,1,1)
		
	pyautogui.keyDown(first_hold)
	pyautogui.keyDown(second_hold)
	time.sleep(1)
	for a in range(0,clicks):
		pyautogui.press(button_click)
	pyautogui.keyUp(first_hold)
	pyautogui.keyUp(second_hold)
	time.sleep(wait)

def htb2(inp):
	#htb2([])
	first_hold = inp[0] 
	second_hold = inp[1]
	button_click = inp[2]
	clicks = inp[3]
	wait = inp[4]
	pyautogui.keyDown(first_hold)
	pyautogui.keyDown(second_hold)
	time.sleep(1)
	for a in range(0,clicks):
		pyautogui.press(button_click)
	pyautogui.keyUp(first_hold)
	pyautogui.keyUp(second_hold)
	time.sleep(wait)






	"""
def rea4([fil,'txt']):
	#fi =fi
	#print "fi",
	fi = fi+".txt"
	fi = open(fi, "r")
	co = fi.read()
	#print co
	return co
	"""

#brs = rea4(["bro",'txt'])[0:3]
brs = "chr"

#brs = rea4(["bro",'txt'])[0:3]


def matn(inp):
	spe = inp[0]
	ray = inp[1]
	ide = inp[2]
	bac = inp[3]
	ret = spe
	for a,val in enumerate(ray):
		print (val[ide])
		if val[ide]==spe:
			ret = val[bac]
			#return val[bac]
			break
	return ret


def upd(inp):
	num = inp[0]
	#num = num
	#num = 2
	fol = os. getcwd()
	if num=="dri":
		fol3 = fol[0:fol.find(":")+2]
	if num!="dri":
		fol2 = fol[::-1]
		beg = 0
		fin = 0

		for a in range(0,num):
			fin = fol2.find("\\",beg)
			beg = fin+1
		fol3 = fol[0:len(fol)-fin]
	return fol3
	#fol = upd([1])


def tes(inp):
	#fol = "B:\\Users\\-\\Downloads"
	fol = upd([1])+"Downloads"
	lis = os.listdir(fol)
	rep = "yt5s.com - "
	for a, val in enumerate(lis):
		if rep in val:
			old = fol+"\\"+val
			nam = val.replace(rep,"")
			new = fol+"\\"+nam
			os.rename(old, new)
	#sye()
	ray = []
	ray.append(["fir",7,"firefox.lnk"])
	ray.append(["chr",15,"chrome.lnk"])
	fol = os.getcwd()
	#fol = fol.replace("0c","Downloads")
	lis = os.listdir(fol)
	app = ".txt"
	for a,val in enumerate(ray):
		val2 = val[0]
		fin = val2+app
		if fin not in lis:
			new = open(fin,"w")
			new.write("")
			new.close() 
	che = inp[0]
	tex = rea4([che,'txt'])
	man = len(tex)
	if man==0:
		lau = mat2([ray,che,0,2])
		wai = mat2([ray,che,0,1])
		sw(lau,wai)
		hod3(["alt","f4",1,0])
		from datetime import datetime
		tim = datetime.now()
		cur = tim.strftime("%H:%M:%S")
		fin = che+".txt"
		new = open(fin,"w")
		new.write(cur)
		new.close()

def opt2(inp):
	#opt2([urr,0])
	ray = inp[0]
	tas = inp[1]
	for a,val in enumerate(ray):
		ray[a] = val
	ray2 = []
	new = []
	nex = 6
	sta = 0
	num = sta
	new = []
	for a,val in enumerate(ray):
		new.append(val)
		num = num+1
		if num==nex or a==len(ray)-1:
			ray2.append(new)
			new = []
			num = sta
	for a,val in enumerate(ray2):
		opt([val,1,tas])
		pgd([len(val),1])
		if tas==1:
			tas=0
	pgu(len(ray),1)		
	print (ray2)


def opt4(inp):
	#opt4([url])
	ray = inp[0]
	fil = "opt7.html"
	tex1 = rea([fil])
	print(ray)
	print(len(ray))
	#sye()
	for a,val in enumerate(ray):
		rep = ""
		for b,valb in enumerate(val):
			print("valb",valb)
			rep = rep+"ray.push(\""+valb+"\");"+"\n\t\t"
		tex2 = tex1.replace("-----",rep)
		out = "opt4.html"
		wri2([out,tex2])
		sw(out,1)
		cf_este3(["open",0,1])
		hod3(["ctrl","pageup",1,1])
		cf4(1,0)
		hod3(["ctrl","9",1,1])



def opt3(inp):
	#opt3([urls])
	ray = inp
	for a,val in enumerate(ray):
		if a==0:
			alt(1,1)
		opt2([val,0])

def alt(cl,wa):
	#alt(1,1)
	hod3(["alt","tab",cl,wa])

def alt3(inp):
	#alt3([1,1])
	cl = inp[0]
	wa = inp[1]
	hod3(["alt","tab",cl,wa])

def cop(inp):
	#
	ray = inp[0]
	num = inp[1]
	var = "copv"
	if num==0:
		f= open(var+".txt","w")
		f.write("0")
		f.close() 
	num = int(rea4([var,'txt']))
	#sys.exit()
	fil = "cop.txt"
	if num==0:
		tex = ""
		for a,val in enumerate(ray):
			tex = tex+val+"\n"
		f= open(fil,"w")
		f.write(tex)
		f.close() 
		sw("cop.txt",1)
	if num>0:
		alt(1,1)
	hod3(["shift","down",1,0])
	hod3(["ctrl","c",1,0])
	but(["down","up"],0,0)
	#but(["down"],0,0)
	if num<len(ray)-1: 
		hod3(["alt","tab",1,0])
	if num==len(ray)-1: 
		af4(1,1)
	hod3(["ctrl","v",1,0])
	print ("num1",num)
	num = num+1
	print ("num2",num)
	f= open(var+".txt","w")
	f.write(str(num))
	f.close() 
	return num

def cop3(inp):
	#cop3(["email"])
	tex = inp[0]
	nam = "cop3.txt"
	wri2([nam,tex])
	sw(nam,1)
	hod3(["ctrl","a",1,0])
	hod3(["ctrl","c",1,0])
	af4(1,1)
	hod3(["ctrl","v",1,0])


def cop2(inp):
	#cop2[ray]
	ray = inp[0]
	inc = int(inp[1])
	tex = ""
	if inc==0:
		for a,val in enumerate(ray):
			if type(val)==str:
				tex = tex+val+"\n"
		wrt2([tex,"cop","txt"])
		sw("cop.txt",2)
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		but(["down","up"],0,0)
	if inc>0:
		alt(1,1)
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		but(["down","up"],0,0)
	alt(1,1)
	hod3(["ctrl","v",1,0])
	inc = inc+1
	pri(tex)
	return inc
	#sye()

	

def wri(to_write,wait):
	#wri("text")
	pyautogui.write(to_write)
	time.sleep(wait)

def pgu(num,wait):
	#for a in range(0,clicks):
	hod3(["ctrl","pageup",num-1,wait])


def af4(cl,wa):
	hod3(["alt","f4",cl,wa])


def but(ray,between_wait,end_wait):
	for a in range(0,len(ray)):
		pyautogui.press(ray[a])
		time.sleep(between_wait)
	time.sleep(end_wait)



def but3(inp):
	#but3(["esc","tab"],0,0)
	ray = inp[0]
	between_wait = inp[1]
	end_wait = inp[2]	

	for a in range(0,len(ray)):
		pyautogui.press(ray[a])
		time.sleep(between_wait)
	time.sleep(end_wait)

def sto(inp):
	#sto([url2])
	url2 = inp[0]
	sff([])
	ini([url2,3])
	olo([])
	"""
	new.append([sff,[]])
	new.append([ini,["naw",3]])
	new.append([olo,[]])
	"""
def ini(inp):
	#ini([fil,3])
	urll = inp[0]
	print("naw")
	print(urll)
	print("naw")

	fol = cuf([])
	nam = "num.html"
	
	htm = ""
	for a,val in enumerate(urll):
		htm = htm+"x.push(\""+val+"\");\n"	
	opt6 = rea4(["opt5","txt"])
	opt6 = opt6.replace("-----",htm)
	wrt2([opt6,"num","html"])
	print(htm)

	dri = os.getcwd()
	dri = dri[0:dri.find("\\")]	

	fil = dri+"\\Users\\-\\"+fol+"\\"+nam
	wai = inp[1]
	os.startfile(fil)
	time.sleep(wai)

def olo(inp):
	#olo([])
	alt(1,0)
	alt(1,0)
	cf_este3(["open",0,2])
	hod3(["ctrl","pageup",1,1])
	cf4(1,0)

def sw(l,w):
	#sw(fil,1)
	os.startfile(l)
	time.sleep(w)

def sw2(inp):
	#sw2(["wor.exe",1])
	fil = inp[0]
	wai = inp[1]
	os.startfile(fil)
	time.sleep(wai)

def hod(first,second,clicks,wait):
	#hod3(["alt","f4",1,0])
	pyautogui.keyDown(first)
	for a in range(0,clicks):
		pyautogui.press(second)
	pyautogui.keyUp(first)
	time.sleep(wait) 


def cot(inp):
	ray = inp[0]
	tab = inp[1]
	wai = inp[2]
	hod3(["alt","tab",tab,0])
	tex = ""
	for a,val in enumerate(ray):
		tex = tex+val+"\n"
	fil = "cot.txt"
	f= open(fil,"w")
	f.write(tex)
	f.close() 
	sw(fil,1)
	for a,val in enumerate(ray):
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		but(["down","up"],0,0)
		if a==len(ray)-1: 
			af4(1,0)
		if a<len(ray)-1: 
			hod3(["alt","tab",1,0])
		hod3(["ctrl","t",1,0])
		hod3(["ctrl","v",1,0])
		key([["enter",1,0,0]])
		if a<len(ray)-1: 
			hod3(["alt","tab",1,0])
	pgu(len(ray),1)		

def aln(inp):
	cl = inp[0]
	wa = inp[1]
	hod3(["alt","tab",cl,wa])


def whi(inq):
	#whi([inq])
	#whi(["input values = "])
	ray = []
	qui = 0
	tex = inq
	add = " = "
	if add not in tex:
		tex = tex+add 
	while(qui<1):
		inp = input(tex)	
		if len(inp)==0:
			break
		ray.append(inp)
	return ray

def whi2(inp):
	#whi([inp])
	ask = inp[0]
	ray = []
	qui = 0
	while(qui<1):
		bac =str(input(ask+" = "))
		if len(bac)==0:
			break
		ray.append(bac)
	return ray


def whi3(inp):
	#whi3(["depart = "])
	inq = inp[0]
	rep = " = "
	if rep not in inq:
		inq = inq+rep
	end = 0
	bac = []
	while(end<1):
		ask = str(input(inq))
		if len(ask)==0:
			break
		if len(ask)>0:
			bac.append(ask)
	return bac

def whi4(inp):
	#inputs = whi4(inputs)
	#inputs = whi4([inputs])
	inputs = inp[0]
	back = []
	for a,val in enumerate(inputs):
		if " = " not in val:
			val = val+" = "
		answer = input(val)
		back.append(answer)
	return back

def cf(inp):
	#cf(["continue"])
	text = inp[0]
	hod3(["ctrl","f",1,1])
	wri(text,1)

def cf5(inp):
	#cf5(["continue",0,1])
	text = inp[0]
	wai1 = inp[1]
	wai2 = inp[1]
	wai3 = inp[1]
	hod3(["ctrl","f",1,wai1])
	wri(text,wai2)




def cf_etste3(inp):
	#cf_etste3(["sign in",2,0])
	tex = inp[0]
	ent = inp[1]
	wai = inp[2]
	hod3(["ctrl","f",1,1])
	wri(tex,1)
	key([["enter",ent,0,1]])
	key([["esc",1,0,0],["tab",1,0,1]])
	hod3(["shift","tab",1,0])
	but(["enter"],0,wai)


def key(ray):
	#key([["down",5,0,0],["enter",1,0,0]])
	for a,val in enumerate(ray):
		button = val[0]
		clicks = val[1]
		#try:
		but = []
		print ("val[1]",clicks)
		for b in range(0,clicks):
			but.append(button)
		print (but)
		but3([[but],val[2],val[3]])


def cf2(inp):
	#cf2(["sea",1])
	sea = inp[0]
	wai = inp[1]
	hod3(["ctrl","f",1,1])
	wri(sea,wai)


def cf_et5(inp):
	text = inp[0]
	wai1=  inp[1]
	wai2 = inp[2]
	hod3(["ctrl","f",1,wai1])
	wri(text,1)
	#but(["esc","tab"],1,wai2)
	but(["esc","tab"],0,wai2)


def cf_et3(inp):
	tex = inp[0]
	wai1 = inp[1]
	wai2 = inp[2]
	hod3(["ctrl","f",1,1])
	wri(tex,wai1)
	#but(["esc"],0,0)
	but3([["esc","tab"],0,wai2])


def cf_este3(inp):
	#cf_este3(["google",0,0])
	text = inp[0]
	wai1 = inp[1]
	wai2 = inp[2]
	hod3(["ctrl","f",1,1])
	wri(text,1)
	but(["esc"],0,0)
	hod3(["shift","tab",1,0])
	but(["enter"],wai1,wai2)

def cf_estt(inp):
	#cf_estt(["jul",0,0])
	text = inp[0]
	wai1 = inp[1]
	wai2 = inp[2]
	hod3(["ctrl","f",1,1])
	wri(text,1)
	but(["esc"],0,0)
	hod3(["shift","tab",1,0])
	but(["tab"],wai1,wai2)

def cf_est(inp):
	#cf_est["forgot",1,1]
	text = inp[0]
	wai1 = inp[1]
	wai2 = inp[2]
	hod3(["ctrl","f",1,1])
	wri(text,1)
	but(["esc"],0,0)
	hod3(["shift","tab",1,0])



def cf_etm3(inp):
	tex = inp[0]
	tab = inp[1]
	hod3(["ctrl","f",1,1])
	wri(tex,1)
	but3(["esc","tab"],0,0)
	hod3(["shift","tab",1,0])


def cf_est3(inp):
	text = inp[0]
	#sc = inp[1]
	key = inp[1]
	wai2 = inp[2]
	hod3(["ctrl","f",1,1])
	wri(text,4)
	but(["esc"],1,1)
	hod3(["shift","tab",key,wai2])


def cf3(inp):
	sea = inp[0]
	wai = inp[1]
	hod3(["ctrl","f",1,1])
	wri(sea,wai)


def gom(ray):
		#ray = ray

	loc = fol(["where"])
	whe = loc[0]

	url = "https://www.google.com/maps/place/"+whe
	url= [url]
	"""
	if type(url)==str:
		url = [url]
		"""
	#u
	#url = []
	#url.append("https://www.google.com/maps/place/"+whe)
	nu(url,1)
	



def sta(inp):
	#l = inp[1]
	#w = inp[0]
	loc = inp[0]
	os.startfile(loc)
	#time.sleep(w)	


def cf_ee5(inp):
	text = inp[0]
	wai1 = inp[1]
	wai2 = inp[2]
	hod3(["ctrl","f",1,1])
	wri(text,1)
	but(["esc","enter"],wai1,wai2)
	#cf_ee5(["this week",0,2])


def cf_etst(inp):
	text = inp[0]
	hod3(["ctrl","f",1,1])
	wri(text,1)
	but(["esc","tab"],0,0)
	hod3(["shift","tab",1,0])
	#but(["enter"],0,0)


def sle(inp):
	#sle([5])
	wai = inp[0]
	#time.sleep(wai)
	for a in range(0,wai):
		#print (a+1,wai,"waiting...",end=" ")
		print (a+1,wai,"waiting...")
		time.sleep(1)




def mot(inp):
	print (inp,len(inp))
	pos = pyautogui.position()
	print (pos)
	def mat(inp):
		spe = inp[0]
		ray = inp[1]
		ide = inp[2]
		bac = inp[3]
		for a,val in enumerate(ray):
			if val[ide]==spe:
				return val[bac]
				break
	print (inp)
	coo = []
	coo.append(["ru",1000,725,360])
	coo.append(["rd",1000,0,360])
	coo.append(["lu",350,725,360])
	coo.append(["ld",350,0,360])
	
	#ope = "fir"
	#sof = "chr"
	
	#sof = rea4(["bro",'txt'])[0:3]
	cor = []
	cor.append(["fir","p","firefox.exe"])
	cor.append(["chr","n","chrome.exe"])
	
	let = matn([brs,cor,0,1])
	fil = matn([brs,cor,0,2])
	
	print (let,fil)
	

	for a,val in enumerate(inp):
		spe = val
		new = ""
		if "r" in spe:
			new = "right"
		if "l" in spe:
			new = "left"
		#alt(1,0)
		hod3(["alt","space",1,0])
		key([["down",5,0,0],["enter",1,0,0]])
		hod3(["win",new,1,0])
		
		#key([["down",2,0,0],["enter",1,0,0],["left",1,0,0],["up",1,0,0],["enter",1,0,0]])
		pox = mat([val,coo,0,1])
		sta = mat([val,coo,0,2])
		end = mat([val,coo,0,3])
		pyautogui.moveTo(pox, sta)	
		pyautogui.mouseDown()
		pyautogui.moveTo(pox, end)	
		pyautogui.mouseUp()


def alr(inp):
	inp = ["fir"]
	tes(inp)


def hod3(inp):
	#hod3(["ctrl","pagedown",1,1])
	first = inp[0]
	second = inp[1]
	clicks = inp[2]
	wait = inp[3]

	pyautogui.keyDown(first)
	for a in range(0,clicks):
		pyautogui.press(second)
	pyautogui.keyUp(first)
	time.sleep(wait) 


def ne72(inp):
	print (inp,len(inp))
	pos = pyautogui.position()
	print (pos)
	def mat(inp):
		spe = inp[0]
		ray = inp[1]
		ide = inp[2]
		bac = inp[3]
		for a,val in enumerate(ray):
			if val[ide]==spe:
				return val[bac]
				break
	print (inp)
	coo = []
	coo.append(["ru",1000,725,360])
	coo.append(["rd",1000,0,360])
	coo.append(["lu",350,725,360])
	coo.append(["ld",350,0,360])
	
	#ope = "fir"
	#sof = "chr"
	
	#sof = rea4(["bro",'txt'])[0:3]
	cor = []
	cor.append(["fir","p","firefox.exe"])
	cor.append(["chr","n","chrome.exe"])
	
	let = matn([brs,cor,0,1])
	fil = matn([brs,cor,0,2])

	wai = 1

	swi = 1
	if swi==1:
		sel = "fir"
		if "fir" in sel:
			#fil = "B:\\Program Files\\Mozilla Firefox\\firefox.exe"
			#fil = "B:\\Program Files (x86)\\Mozilla Firefox\\firefox.exe"

			#fil = upd(["dri"])+"Program Files (x86)\\Mozilla Firefox\\firefox.exe"



			#fil = upd(["dri"])+"Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
			fil = upd(["dri"])+"Program Files (x86)\\Mozilla Firefox\\firefox.exe"
			let = ""
			if "chrome" in fil:
				let = "n"

				new = open("bro.txt","w")
				new.write("chrome")
				new.close()


			if "firefox" in fil:
				let = "p"	

				new = open("bro.txt","w")
				new.write("firefox")
				new.close()

			wai = 3

	print (let,fil)

	for a,val in enumerate(inp):
		spe = val
		sw(fil,wai)
		htb("ctrl","shift",let,1,1)
		alt(1,0)
		af4(1,0)
		new = ""
		print ("spe",spe)
		if "r" in spe:
			new = "right"
		if "l" in spe:
			new = "left"
		hod3(["win",new,1,1])


def net(inp):
	ray = []
	if inp[0]==1:
		#ray = []
		ray.append("bro")
		ray.append("res")
		ray = fol(ray)
	if inp[0]!=1:
		ray = inp[1]
	inp = ray
	print ("inp",inp)
	#sys.exit()
	cor = inp[0]
	com = inp[1]
	fir = []
	#fir.append("firefox.lnk")
	
	fir.append("C:\\Program Files\\Mozilla Firefox\\firefox.exe")
	fir.append("p")
	fir.append(4)
	goo = []
	goo.append("chrome.exe")
	goo.append("n")
	goo.append(5)
	ray = []
	ray.append(["fir",fir])
	ray.append(["chr",goo])
	print 
	print ("cor",cor)
	for a,val in enumerate(ray):
		if val[0]==cor:
			"""
			loc = val[1][0]
			key = val[1][1]
			red = val[1][2]
			"""
			ini([1,val[1][0]])
			twd(["ctrl","shift",val[1][1],1,1])
			alt(1,1)
			af4(1,1)		
			res([com,val[1][2]])		
			break


def opt(inp):

	bro = rea4(["bro","txt"])

	sof = rea4(["bro",'txt'])[0:3]
	ope = []
	ope.append(["fir",cf_ete,"open"])
	ope.append(["chr",cf_ee,"open"])
	cor = matn([sof,ope,0,1])
	var = matn([sof,ope,0,2])
	ray = inp[0]
	tab = inp[1]
	tac = inp[2]
	tex = ""
	print (ray)
	for a,val in enumerate(ray):
		if "http" not in val:
			if "www." not in val:
				val = "www."+val
			ray[a] = "https://"+val
	if "firefox" in bro:
		ray = ray[::-1]
	for a,val in enumerate(ray):
		tex = tex+"x.push(\""+val+"\");\n\t\t"
	print (tex)
	rep = "___rep___"
	con = rea4(["opb",'txt'])
	con = con.replace(rep,tex)
	fil = "opt.html"
	print (fil)
	new = open(fil,"w")
	new.write(con)
	new.close()
	print (con)
	loc = upd([0])
	loc = loc+"\\"+fil
	ope = [loc]
	cot([ope,tac,1])
	
	
	if "chrome" in bro:
		cf_ee("open")
		time.sleep(1)
		pgu(len(ray)+1,1)		
		hod3(["ctrl","f4",1,0])


	if "firefox" in bro:
		cf_ete("open")
		time.sleep(1)
		pgu(2,1)		
		hod3(["ctrl","f4",1,0])


def var(inp):
	ray = []
	for a,val in enumerate(inp):
		bac = input(val+" = ")
		ray.append(bac)
	return ray


def cap(tex):
	return tex.capitalize()

def upp(tex):
	#tex = tex.upper()
	return tex.upper()


def num4(inp):
	#num4([url,2])
	ray = inp[0]
	if type(ray)==str:
		ray = [ray]
	wai = inp[1]
	#num4([[url],wai])
	tex = ""
	for a,val in enumerate(ray):
		tex = tex+val+"\n"
	fil = "num.txt"
	wrt2([tex,fil,""])
	#wai..wait?? inp[1]??
	alt(1,1)
	subprocess.Popen(['notepad.exe', fil])
	sle([wai])
	for a,val in enumerate(ray):
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		but(["down","up"],0,0)
		hod3(["alt","tab",1,0])		
		hod3(["ctrl","t",1,0])
		hod3(["ctrl","v",1,0])
		but(["enter"],0,0)
		hod3(["alt","tab",1,0])
		if a==len(ray)-1:
			hod3(["alt","f4",1,0])

def num5(inp):
	#url
	#num5([urls,2])
	urls = inp[0]
	if type(urls)==str:
		urls = [urls]
	wait = inp[1]
	for a,val in enumerate(urls):
		#if a==0:
			#alt(1,1)
		save = "num5.txt"
		text = val+"\n"
		wri2([save,text])
		sw2([save,1])
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		af4(1,1)
		hod3(["ctrl","t",1,0])
		hod3(["ctrl","v",1,0])
		but(["enter"],0,0)	

def num6(inp):
	#url
	#num6([urls,wait time,yes or no alt tab, 1 for yes, 0 for no])
	#num6([urls,2,1])
	urls = inp[0]
	if type(urls)==str:
		urls = [urls]
	wait = inp[1]
	alt_tab = inp[2]
	for a,val in enumerate(urls):
		if a==0:
			if alt_tab==1:
				alt(1,1)
		save = "num5.txt"
		text = val+"\n"
		wri2([save,text])
		sw2([save,1])
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		af4(1,1)
		hod3(["ctrl","t",1,0])
		hod3(["ctrl","v",1,0])
		but(["enter"],0,0)	
	time.sleep(wait)	

def nu(ra,at):
	lt(ra)
	hod3(["alt","tab",at,0])
	na = os.path.basename(__file__)
	na = na.replace(".py",".txt")
	lo = os.getcwd()+"\\"+na
	#subprocess.Popen(['notepad.exe', lo])
	subprocess.Popen(['notepad.exe', lo])
	time.sleep(1)
	lo = lo.replace(".txt","")
	co = rt(lo)
	le = co.count("\n")
	las = le-1
	if las<0:
		las=0
	for a in range(0,le):
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		but(["down","up"],0,0)
		hod3(["alt","tab",1,0])		
		hod3(["ctrl","t",1,0])
		hod3(["ctrl","v",1,0])
		but(["enter"],0,0)
		hod3(["alt","tab",1,0])
		if a==las:
			af4(1,1)

def num(inp):
	#num([url])
	url = inp[0]
	if type(url) is str:
		url = [url] 
	tex = ""
	for a,val in enumerate(url):
		tex = tex+val+"\n"
	nam = "num.txt"
	wri2([nam,tex])
	alt(1,1)
	#subprocess.Popen(['notepad.exe', nam])
	sw(nam,2)
	for a,val in enumerate(url):
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		but(["down","up"],0,0)
		if a<len(url)-1:
			hod3(["alt","tab",1,1])	
		if a==len(url)-1:
			af4(1,1)	
		hod3(["ctrl","t",1,0])
		hod3(["ctrl","v",1,0])
		but(["enter"],0,0)
		if a<len(url)-1:
			hod3(["alt","tab",1,1])	

def lt(ra):
	te = ""
	for a in range(0,len(ra)):
		te = te+ra[a]+"\n"
	na = os.path.basename(__file__)
	na = na.replace(".py","")
	f= open(na+".txt","w")
	f.write(te)
	f.close() 




def dis(inp):
	inp = inp
	for a,val in enumerate(inp):
		print (a,val)


def rt(fi):
	fi =fi+".txt"
	#print "fi",
	fi = open(fi, "r")
	co = fi.read()
	#print co
	return co

def inr3(rer):
	for a in range(0,len(rer)):
		rer[a] = [rer[a]]

	for a in range(0,len(rer)):
		
		pro = rer[a][0]+" = "
		rer[a].append(input(pro))
	return rer

def at(cl,wa):
	hod3(["alt","tab",cl,wa])

def ud(st):
	st = st[::-1]
	fi = st.find("\\")
	st = st[fi+1:len(st)]
	st = st[::-1]
	print(st)
	return st

def mat2(inp):
	#cor = mat2([ray,"fir",0,2])
	ray = inp[0]
	sea = inp[1]
	ide = inp[2]
	bac = inp[3]
	for a,val in enumerate(ray):
		if val[ide]==sea:
			return val[bac]

def sye():
	sys.exit()

def upd2(inp):
	cwd = os.getcwd()
	qui = 0
	sta = 0
	bac = ""
	las = 0
	while(qui<1):
		loc = cwd.find("\\",sta)
		print (loc)
		if loc<0:
			bac = cwd[0:las]+"\\"
			break
		sta = loc+1
		las = loc
	print (bac)
	return (bac)
	#rev = cwd[::-1]


def upd3(inp):
	cwd = os.getcwd()
	rev = cwd[::-1]
	fin = rev.find("\\")
	#rev2 = rev[0:fin]
	bac = cwd[0:len(cwd)-fin]
	print (bac)


def twd(inp):
	first_hold = inp[0]
	second_hold = inp[1]
	button_click = inp[2]
	clicks = inp[3]
	wait = inp[4]

	pyautogui.keyDown(first_hold)
	pyautogui.keyDown(second_hold)
	time.sleep(1)
	for a in range(0,clicks):
		pyautogui.press(button_click)
	pyautogui.keyUp(first_hold)
	pyautogui.keyUp(second_hold)
	time.sleep(wait)


def pri(inp):
	array = inp
	print("--------------------list below-----------------------")
	try:
		for a,val in enumerate(array):
			print(a," = ",val)
	except:
		#fail = "fail"
		print("don't work! type of array = "+str(type(array)))

		"""
	if type(inp)==list:
		for a,val in enumerate(inp):
			print(str(a)+" = "+str(val))
	else:
		print(str(inp))
		"""

def pri2(inp):
	ray = inp[0]
	print ("")
	for a,val in enumerate(ray):
		print (a,len(val),val)

def pri3(inp):
	for a,val in enumerate(inp):
		if type(val) is str:
			print (val)
		if type(val) is list:
			for b,val2 in enumerate(val):
				print (b,val2)

	ray = inp[0]
	print ("")
	for a,val in enumerate(ray):
		print (a,len(val),val)


def nex(inp):
	#nex([tex,"ide",sta])
	#nex([tex,"ide",0])
	tex = inp[0]
	ide = inp[1]
	sta = inp[2]
	#rep = inp[3]
	ray = []
	qui = 0
	end = 0
	while(qui<1):
		fin = tex.find(ide,sta)
		if fin<0:
			#break
			fin = len(tex)
			end = 1
		lin = tex[sta:fin]
		ray.append(lin)
		#sta = fin+1
		sta = fin+len(ide)
		if end==1:
			break
		#for a,val in enumerate(rep):
		#	lin = lin.replace(val,"")
	return ray

def nex3(inp):
	#nex3([fil/tex,ide1,ide2])
	fil = inp[0]
	ide1 = inp[1]
	ide2 = inp[2]
	#fil = "red2.txt"
	if type(fil)==str:
		tex = fil
	else:
		tex = rea([fil])
	#print("tex", tex)
	cou = tex.count("\n-")
	cou2 = tex.count("\n\n")
	#print(cou,cou2)
	tex = tex
	end = 0
	#ide1 = "\n-"
	#ide2 = "\n\n"
	beg = 0
	inc = 0
	bac = []
	while(end<1):
		fin = tex.find(ide1,beg)
		if fin<0:
			break
		sto = tex.find(ide2,fin)
		lin = tex[fin:sto].replace(ide1,"")
		#if len(lin)==0:
		#	break
		bac.append(lin)
		inc = inc+1
		#print(inc,lin)
		beg = fin+1
	return bac

def nex4(inp):
	#back = nex4([text,iden1,iden2])
	#back = nex4([text,"\n","\n"])
	text = inp[0]
	iden1 = inp[1]
	iden2 = inp[2]
	back = []
	quit = 0
	while(quit<1):
		if len(back)==0:
			loc1 = 0
		if len(back)>0:
			loc1 = text.find(iden1,loc2)
		loc2 = text.find(iden2,loc1+1)
		if loc2<0:
			loc2=len(text)
			quit = 1
		app = text[loc1:loc2]
		back.append(app)
	#pri(back)
	return back

def nex2(inp):
	#got rid of 0 on end
	#nex2([tex,"ide","</span>"])
	#old way to do it
	#nex2([tex,"ide","end",sta])
	#nex2([tex,"ide","end",0])
	#nex2([tex,"ide","</span>",0])
	tex = inp[0]
	ide = inp[1]
	end = inp[2]
	#sta = inp[3]
	#rep = inp[3]
	ray = []
	qui = 0
	sta = 0
	while(qui<1):
		fin = tex.find(ide,sta)
		print ("fin",fin)
		if fin<0:
			break
		#print (fin)
		sto = tex.find(end,fin)
		if sto<0:
			end = ">/table"
			sto = tex.find(end,fin)

		print ("end",end)
		print ("sto",sto)
		if sto<fin:
			break
		lin = tex[fin:sto]
		ray.append(lin)
		sta = sto+1
		#print(fin,sto,lin,sta)
	return ray


def rep(inp):
	#day2 = rep([day2,[[", ",""]]])
		
	ray = inp[0]
	lis = inp[1]
	if type(ray)==str:
		for a,val in enumerate(lis):
			ray = ray.replace(val[0],val[1])
	"""
	gon = []
	gon.append(["M","000000"])
	gon.append(["B","000000000"])
	lis2 = rep([lis,gon])
	"""
	if type(ray)==list:
		for a,val in enumerate(ray):
			for b,val2 in enumerate(lis):
				val = val.replace(val2[0],val2[1])
				ray[a] = val
	return ray

def xls(inp):
	#xls(["fil","she",xls")
	nam = inp[0]
	she = inp[1]
	typ=inp[2]
	if "."+typ not in nam:
		nam = nam+"."+typ
	fil = nam
	print(fil)
	import xlrd
	loa = xlrd.open_workbook(fil)
	print(type(loa))
	loa2 = loa[she]
	#pri(loa2)
	row = loa2.nrows
	col = loa2.ncols
	pri([row,col])
	val = loa2.cell_value(0,0)
	pri(val)
	ray=[]
	for a in range(0,row):
		app = []
		for b in range(0,col):
			try:
				spe = loa2.cell_value(a,b)
			except:
				continue
			app.append(spe)
		ray.append(app)
	#pri(ray)
	return ray
	

def xlr(inp):
	#xlr(["fil","xlsx")
	nam = inp[0]
	typ=inp[1]
	if "."+typ not in nam:
		nam = nam+"."+typ
	fil = nam
	#fil = nam+"."+typ
	from openpyxl import load_workbook
	wb = load_workbook(filename = fil)
	sheet = wb['Sheet1']
	row = sheet.max_row
	col = sheet.max_column
	print (row,col)
	ray=[]
	for a in range(0,row):
		new = []
		for b in range(0,col):
			let = chr(ord('@')+(b+1))
			cel = let+str(a+1)
			new.append(sheet[cel].value)	
		ray.append(new)
	new = []
	for a,val in enumerate(ray):
		app = []
		for b,val2 in enumerate(val):
			if "NoneType" in str(type(val2)):
				continue
			app.append(val2)
		new.append(app)
	ray = new
	return ray

def wrx(inp):
	#wrx(["fil","xls",ray]) 
	# fil = name, "xls" = type, ray = array to be written
	nam = inp[0]
	typ = inp[1]
	ray = inp[2]
	sav = nam+'.'+typ
	workbook = xlsxwriter.Workbook(sav)
	worksheet = workbook.add_worksheet()
	for a,val in enumerate(ray):
		for b,val2 in enumerate(val):
			worksheet.write(a,b,val2)
	workbook.close()
	print("file = "+sav)

def wrx2(inp):
	#wrx(["fil","xls",ray]) 
	# fil = name, "xls" = type, ray = array to be written
	nam = inp[0]
	typ = inp[1]
	ray = inp[2]
	wid = []
	for a in range(0,len(ray[0])):
		wid.append([])
	workbook = xlsxwriter.Workbook(nam+'.'+typ)
	worksheet = workbook.add_worksheet()
	for a,val in enumerate(ray):
		for b,val2 in enumerate(val):
			worksheet.write(a,b,val2)
			try:
				wid[b].append(len(str(val2)))
			except:
				continue
	print ("len", len(wid))
	print (wid)
	for a in range(0,len(wid)):
		worksheet.set_column(a,a,max(wid[a]))
	workbook.close()

def xls2(inp):
	#write/create xls file
	#xls2([ray,"fil"])
	ray = inp[0]
	sav = inp[1]
	if ".xls" not in sav:
		sav = sav+".xls"
	import xlwt
	from xlwt import Workbook
	wb = Workbook()
	ws = wb.add_sheet('Sheet1')
	for a,vala in enumerate(ray):
		for b,valb in enumerate(vala):
			ws.write(a,b,valb)
	wb.save(sav)

def xlw(inp):
	#xlw(["fil","xls",ray]) 
	# fil = name, "xls" = type, ray = array to be written
	nam = inp[0]
	typ = inp[1]
	ray = inp[2]
	wid = []
	try:
		for a in range(0,len(ray[0])):
			wid.append([])
	except:
		naw = "naw"
	out = nam+'.'+typ
	workbook = xlsxwriter.Workbook(out)
	worksheet = workbook.add_worksheet()
	for a,val in enumerate(ray):
		for b,val2 in enumerate(val):
			worksheet.write(a,b,val2)
			try:
				wid[b].append(len(str(val2)))
			except:
				continue
	#print ("len", len(wid))
	#print (wid)
	for a in range(0,len(wid)):
		worksheet.set_column(a,a,max(wid[a]))
	workbook.close()
	print(out)

def xlw2(inp):
	#xlw2(["fil","xls",ray,wid])
	nam = inp[0]
	typ = inp[1]
	ray = inp[2]
	wid = inp[3]
	workbook = xlsxwriter.Workbook(nam+'.'+typ)
	worksheet = workbook.add_worksheet()
	for a,val in enumerate(ray):
		for b,val2 in enumerate(val):
			worksheet.write(a,b,val2)
	print ("len", len(wid))
	print (wid)
	if len(wid)>0:
		for a in range(0,len(wid)):
			worksheet.set_column(a,a,wid[a])
		workbook.close()


def hea2(inp):
	for a,val in enumerate(inp):
		inp[a] = [val]
	return inp

def mat(inp):
	#mat([num,ray])
	#mat(["1",ray])
	sea = inp[0]
	ray = inp[1]
	for a,val in enumerate(ray):
		if val[0]==sea:
			bac = val[1]
			break
	return bac

def pro2(inp):
	#cen = pro2([a,len(ray),2])
	inc = inp[0]
	le2 = inp[1]
	af0 = inp[2]
	cen2 = (float(inc)/float(le2))*100
	cen3 = str(dec([cen2,2]))+"%"+" done"
	return cen3

def pro(inp):
	#pro([ray,num,val])
	#pro([ray,a,val])
	naw = "naw"
	ray = inp[0]
	num = inp[1]
	val = inp[2]
	le2 = len(ray)
	#dec([val,2])
	num2 = num+1
	how = (float(num2)/float(le2))*100
	how2 = dec([how,2])
	out = str(how2)+"%"+" done"
	out = out+", "+str(num2)+" done of "+str(le2)
	#out=","+str(le2)+" remaining",
	print(out)

def convert_to_percent(inp):
	#number = convert_to_percent([inp])
	number = inp[0]
	number = float(number)
	number = number-1
	number = number*100
	number = decimal_places([number,2])
	return number


	"""
def pro2(inp):
	#pro2([a,len(array),val])
	inc = inp[0]
	le2 = inp[1]
	val = inp[2]
	val2 = val
	if type(val)==list:
		val2 = val[0]
	out1 = (float(inc)/float(le2))*10000
	out2 = int(out1)
	out3 = str(out2/100)+"%"+", val = "+str(val2)
	print(out3)

def pro3(inp):
	#pro3([tex,a,ray,val])
	sou = inp[0]
	inc = inp[1]
	le2 = inp[2]
	val = inp[3]

	pro2([inc+1,len(le2),val])
	print(str(len(le2)-inc)+" remaining for "+sou+"...")
	"""




def dec(inp):
	#bac = dec([val,2])
	num = inp[0]
	dec = inp[1]
	num = str(num)
	if "e" in num:
		num=0
	else:
		fi = num.find(".")
		num = num[0:fi+dec+1]
	num = float(num)
	return num 

def decimal_places(inp):
	#bac = decimal_places([val,2])
	number = inp[0]
	how_many_decimals = inp[1]
	check = str(number)
	if "e" in check:
		back = 0
	if "e" not in check:
		find_dot = check.find(".")
		if find_dot<0:
			back = number
		if find_dot>=0:
			number2 = check[0:find_dot+how_many_decimals+1]
			back = float(number2)
	return back

def n2d(inp):
	#n2d([val])
	val = inp[0]
	val = ((val-1)*100)
	return val
	

def js2(inp):
	#prh = js2([fil])
	fil = inp[0]
	with open(fil, 'r') as f:
		data = json.load(f)
	return data

def load_json(inp):
	#earnings_history = load_json([file])
	file = inp[0]
	with open(file, 'r') as f:
		data = json.load(f)
	return data


def jsw(inp):
	#jsw([fil,dat])
	#jsw([fil2,dat])
	fil = inp[0]
	dat = inp[1]
	with open(fil,"w") as outfile:
		json.dump(dat,outfile)



def pag(inp):
	#n for no ctrl u, y for yes control u
	#pag(["url",5,"n",location])
	url = inp[0]
	wai = inp[1]	
	u = inp[2]
	fil = inp[3]
	num4([[url],wai])
	if "y" in u:
		hod3(["ctrl","u",1,2])
	hod3(["ctrl","s",1,2])
	hod3(["ctrl","a",1,0])
	#loc = "C:\\Users\\-\\0c\\"+out+".txt"
	#out = fil+".txt"
	#wri(out,2)
	tex = []
	tex.append(fil)
	cop([tex,0])		
	key([["enter",2,1,0]])
	#af4(1,1)
	cf4(1,0)
	cf4(1,0)

def pag2(inp):
	#n forf no ctrl u, y for yes control u
	#pag(["url",5,"n",location])
	url = inp[0]
	wai = inp[1]	
	u = inp[2]
	fil = inp[3]
	num4([[url],wai])
	#if "y" in u:
	#	hod3(["ctrl","u",1,2])
	hod3(["ctrl","a",1,1])
	hod3(["ctrl","c",1,1])
	sw("wordpad.exe",1)
	hod3(["ctrl","v",1,1])
	hod3(["ctrl","s",1,1])
	tex = []
	tex.append(fil)
	cop([tex,0])		
	key([["enter",2,1,0]])
	cf4(1,0)
	cf4(1,0)

def pag3(inp):
	#n for no ctrl u, y for yes control u
	#pag3(["url",5,"n",location])
	url = inp[0]
	wai = inp[1]	
	u = inp[2]
	fil = inp[3]
	num4([[url],wai])
	#if "y" in u:
	#	hod3(["ctrl","u",1,2])
	hod3(["ctrl","s",1,2])
	tex = []
	tex.append(fil)
	cop([tex,0])		
	key([["enter",2,1,0]])
	cf4(1,0)
	alt(1,1)
#	cf4(1,0)

def pag4(inp):
	#n for no ctrl u, y for yes control u
	#pag4(["url",5,"n",location])
	url = inp[0]
	wai = inp[1]	
	u = inp[2]
	fil = inp[3]
	num4([[url],wai])
	#if "y" in u:
	#	hod3(["ctrl","u",1,2])
	pyautogui.moveTo(578, 448)
	pyautogui.click()
	time.sleep(3)
	cf_ee("show more")
	time.sleep(3)
	hod3(["ctrl","a",1,1])
	hod3(["ctrl","c",1,1])
	sw("wordpad.exe",1)
	hod3(["ctrl","v",1,1])
	hod3(["ctrl","s",1,1])
	tex = []
	tex.append(fil)
	cop([tex,0])		
	key([["enter",2,1,0]])
	cf4(1,0)
	cf4(1,0)


def cf4(cl,wa):
	hod3(["ctrl","f4",cl,wa])


def tid(inp):
	#dst = inp[0]
	today = datetime.date.today()
	#day = today.weekday()
	dat2 = datetime.datetime.strptime(dat,'%b %d, %Y')
	las = today-datetime.timedelta(day)
	


def ttr(inp):
	tex = inp[0]
	#print (tex,type(tex),len(tex))
	inf = nex2([tex,"tr>","<tr",0])
	for a,val in enumerate(inf):
		inf[a] = val[0:val.find("</td>")]
	#pri(inf)
	return inf


def num2(inp):
	url = [inp[0]]
	wai = inp[1]
	if len(url)>0:
		#for a,val in enumerate(ray):
		"""
			if len(val)!=1:
				ray[a] = [val]
				"""
		nu(url,1)
		pgu(len(url),1)
		time.sleep(wai)

def num3(inp):
	#num3([])
	import win32api,win32con
	from pynput.keyboard import Key, Controller
	keyboard = Controller()
	def cap():
	    # return 1 if CAPSLOCK is ON
	    #bac = win32api.GetKeyState(win32con.VK_CAPITAL)
	    bac = win32api.GetKeyState(win32con.VK_NUMLOCK)
	    return bac
	val = cap()
	print (type(val))
	print(val)
	if val==1:
		keyboard.press(Key.num_lock)
	#keyboard.press(Key.caps_lock)



def wri3(ray):
	#wri3(["violin78@protonmail.com",2])
	to_write = ray[0]
	wait = ray[1]
	pyautogui.write(to_write)
	time.sleep(wait)

def wri4(inp):
	#wri4([2,"violin78@protonmail.com"])
	wai = inp[0]
	tex = inp[1]
	pyautogui.write(tex)
	time.sleep(wai)

def wri5(inp):
	#wri5(inp)
	#wri5([file_name,file_type,text])
	#wri5(["out","txt",text])
	file_name = inp[0]
	file_type  = inp[1]
	text = inp[2]
	if "." not in file_type:
		file_type="."+file_type
	save = file_name+file_type
	new = open(save,"w")
	new.write(text)
	new.close()


def get(inp):
	#get([xpath])
	loc = inp[0]
	qui = 0
	ind = -1
	mas = []
	ret = []
	while(qui<1):
		ind = ind+1
		#app = []
		#for a,val in enumerate(get):
		loc = val.replace("___",str(ind))
		bac = sel.xpath(loc).getall()
		try:
			bac = bac[0]
		except:
			naw = "naw"
		ret.append(bac)
		#mas.append(app)
		#ind = int(ind)
		if ind>5:
			if len(bac)==0:
				break
	return ret


def xpa(inp):
	fil = inp[0]
	loc = inp[1]
	from parsel import Selector
	f =open(fil,'r',encoding='utf-8')
	doc = f.read()
	f.close()
	sel = Selector(doc)
	qui = 0
	mas = []
	app = []
	num = -1
	while(qui<1):
		num = num+1
		spe = loc.replace("___",str(num))
		#print ("loc",loc)
		bac = sel.xpath(spe).getall()
		bac2 = ""
		if len(bac)>0:
			bac2 = bac[0]
			if len(bac2)>0:
				app.append(bac2)
		#print (num,bac2)
		#app.append(bac)
		if num>5:
			if len(bac)==0:
				break
	#pri(app)
	return app

def ges(inp):
	bac = ""
	for a,val in enumerate(inp):
		bac = bac+val
	return bac


def wai(inp2):
	tem =str(input("click when done loading"))
	at(1,1)


def mad(inp):
	#mad([ope,dayb])
	ope = inp[0]
	dayb = inp[1]
	res = float(int(((ope/dayb)-1)*10000)/100)
	return res

def tki(inp):
	import tkinter as tk
	tex = tk.Tk()
	#tex = tex.decode('utf8')
	try:
		tex2 = tex.clipboard_get()
	except:
		import win32clipboard
		#win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
		win32clipboard.OpenClipboard()
		tex2 = win32clipboard.GetClipboardData().encode('utf-8')
		tex2 = tex2.decode('utf8')
		win32clipboard.CloseClipboard()
	print (tex2)
	print (type(tex2))
	return tex2
	
def fin(inp):
	#lis = fin([main text, begin identifiter, end identifier])
	#lis = fin([tex,"r/","comments\n"])
	tex = inp[0]
	ide1 = inp[1]
	ide2 = inp[2]
	qui =0
	ray = []
	sta = 0
	while(qui<1):
		if type(ide1) is str:
			loc1 = tex.find(ide1,sta)
		if type(ide1) is int:
			loc1 = ide1
		if loc1<0:
			break
		if ide2 ==-1:
			loc2 = len(tex)
		else:
			loc2 = tex.find(ide2,loc1+1)
		if loc2<0:
			break
		mor = 0
		if ide2!=-1:
			mor = len(ide2)

		spe = tex[loc1:loc2+mor]
		ray.append(spe)
		sta = loc2+1
		if type(ide1) is int:
			break
	if len(ray)==1:
		ray = ray[0]
	return ray	

def ski(inp):
	#lis2 = ski([lis,avo])
	lis=inp[0]
	avo=inp[1]
	lis2 = []
	for a,val in enumerate(lis):
		aft = 0
		for b,val2 in enumerate(avo):
			if val2 in val:
				aft = 1
				break
		if aft==1:
			continue
		if aft==0:
			lis2.append(val)
	return lis2

def rep2(inp):
	#lis2 = rep2([lis,new])
	lis2  = []
	lis = inp[0]
	new = inp[1]
	if type(lis)==str:
		lis = [lis]
	for a,val in enumerate(lis):
		val3 = []
		val3.append(val)
		val3 = val3[0]
		for b,val2 in enumerate(new):
			beg = val2[0]
			end = val2[1]
			if type(beg)==str:
				if beg in val3:
					val3 = val3.replace(beg,end)
		lis2.append(val3)
	#pri(lis2)
	if len(lis2)==1:
		lis2 = lis2[0]
	return lis2

def rep3(inp):
	#list = rep3(input list,replacement values)
	#list = rep3(stocks,day_replacement_codes)
	lis = inp[0]
	replacement_values = inp[1]
	for a,val in enumerate(lis):
		for b,valb in enumerate(replacement_values):
			original = valb[0]
			new = valb[1]
			lis[a] = val.replace(original,new)
	return lis


	for a in range(0,len(lis)):
		val = lis[a]
		for b,valb in enumerate(replacement_values):
			original = valb[0]
			new = valb[1]

			if valb in val:
				val = val.replace()

		lis[a] = lis[a].replace()

def abb(inp):
	#val = abb([val])
	#val = abb([val])
	giv = inp[0]
	ray = ["K","M","B"]
	#giv2 = "---"
	giv2 = 0
	if "K" in giv:
		giv2 = giv.replace("K","")
		#giv2 = int(float(giv2))
		giv2 = float(giv2)
		giv2 = giv2/1000000
	if "M" in giv:
		giv2 = giv.replace("M","")
		#giv2 = int(float(giv2))
		giv2 = float(giv2)
		giv2 = giv2/1000
	if "B" in giv:
		giv2 = giv.replace("B","")
		#giv2 = int(float(giv2))
		giv2 = float(giv2)
		giv2 = giv2/1
	return giv2

def abb2(inp):
	#val = abb2([val])
	giv = inp[0]
	ray = ["K","M","B"]
	#giv2 = "---"
	giv2 = 0
	if "K" in giv:
		giv2 = giv.replace("K","")
		#giv2 = int(float(giv2))
		giv2 = float(giv2)
		giv2 = giv2*1000
	if "M" in giv:
		giv2 = giv.replace("M","")
		#giv2 = int(float(giv2))
		giv2 = float(giv2)
		giv2 = giv2*1000000
	if "B" in giv:
		giv2 = giv.replace("B","")
		#giv2 = int(float(giv2))
		giv2 = float(giv2)
		giv2 = giv2*1000000000
	giv2 = int(giv2)
	return giv2

def mkd(inp):
	#mkd(["ear",cwd])
	che = inp[0]
	fol = inp[1]
	lis = os.listdir(fol)
	if che not in lis:
		new = fol+"\\"+che
		os.mkdir(new)

def che(inp):
	#mis = che([lise,nee])
	inf = inp[0]
	hav = inp[1]
	mis = []
	for a,val in enumerate(hav):
		if val not in inf:
			mis.append(val)
	return mis

def cha2(inp):
	#val = cha([val/val2])
	#gap = cha([gap])
	if len(inp)==2:
		bef = inp[0]
		aft = inp[1]
		div = float(aft)/float(bef)
	if len(inp)==1:
		div = inp[0]
	res = (div-1)*100
	return res

def ctu(inp):
	#ctu([urls])
	#ctu([url])
	url = inp[0]
	if type(url)==str:
		url = [url]
	tex = ""
	for a,val in enumerate(url):
		tex = tex+val+"\n"
	fil = "ctu.txt"
	wri2([fil,tex])
	for a,val in enumerate(url):
		if a==0:
			alt(1,1)
			sw(fil,2)
		hod3(["shift","down",1,0])
		hod3(["ctrl","c",1,0])
		key([["down",1,0,0],["up",1,0,0]])
		if a==len(url)-1:
			af4(1,1)
		else:
			alt(1,1)
		hod3(["ctrl","t",1,0])
		hod3(["ctrl","v",1,0])
		key([["enter",1,0,0]])

def opa(inp):
	#opa(["aut3.html",urls])
	#htm = opa([output html file name,list of urls])
	urls = inp[1]
	fil = "opa.txt"
	lis = []
	out = ""
	for a,val in enumerate(urls):
		spe = "x.push(\""+val+"\");\n\t"
		out = out+spe
	tex = rea([fil])
	tex2 = tex.replace("-----",out)
	out2 = inp[0]
	if ".html" not in out2:
		out2 = out2+".html"
	wri2([out2,tex2])

def opa2(inp):
	#opa2([])
	cf_ete2(["open",2])
	#htb2(["ctrl","shift","tab",1,0])
	hod3(["ctrl","pageup",1,2])
	hod3(["ctrl","f4",1,0])

def opa3(inp):
	#opa3(["aut3",urls])
	#opa3(["loc",urls])
	nam = inp[0]
	urls = inp[1]
	fil = nam+".html"
	acc2 = acc([])
	fil_loc = "C:\\Users\\"+acc2+"\\util\\"+fil
	opa([fil,urls])
	#sw2(["firefox.lnk",1])
	sw2([fil_loc,3])
	opa2([])

def voi2(inp):
	#voi2([])
	exec(open("voi.py").read())

def get_tab(inp):
	#ray = get_tab(inp):
	#ray = get_tab([name of txt file,format of file]]):
	#ray = get_tab(["mario",".txt"]]):	
	fil = inp[0]
	form = inp[1]
	fil = fil+"."+form
	text = rea([fil])
	#pri(text)
	back = nex4([text,"\n","\n"])
	for a,val in enumerate(back):
		back[a] = val.replace("\n","")
	pri(back)
	back2 = []
	for a,val in enumerate(back):
		new = []
		new = nex4([val," "," "])
		back2.append(new)
	back3 = []
	for a,val in enumerate(back2):
		if a==0:
			back3.append(val)
			continue
		improvement = []
		for b,valb in enumerate(val):
			check = valb.replace(" ","")
			if len(check)>0:
				#back2[a][b] = int(valb.replace(" ",""))
				#improvement.append(valb.replace(" ",""))
				improvement.append(check)
				print(len(valb))
		back3.append(improvement)
		#back2[a]=val[0].replace(" ","")
	pri(back3)
	back2 = back3
	#sye()
	title = ""
	for a,val in enumerate(back2[0]):
		title = title+val
	back2[0]=[title]
	pri(back2)
	#sye()
	return back2
def transpose(inp):
	#final = transpose([inp])	
	#final = transpose([array to transpose,octave up or down])
	#final = transpose([back,"up"])
	#final = transpose([back,"down"])
	back3 = inp[0]
	#direction = "up" or "down"
	direction = inp[1]
	#drop = "yes"
	#back2 = []
	if direction=="up" or direction=="down":
		back4 = []
		for a,val in enumerate(back3):
			if a==0:
				continue
			app = []
			for b,valb in enumerate(val):
				if len(valb)>0:
					print(a,b,valb)
					#print(valb)
					new = int(valb)
					if direction=="up":
						if new<0:
							if new<-4:
								new = new-4
						if new>0:
							new = new-3
					app.append(new)
			back4.append(app)
	return back4

repetoire = []
repetoire.append([1,"C4"])
repetoire.append([2,"E4"])
repetoire.append([3,"G4"])
repetoire.append([4,"C5"])
repetoire.append([5,"E5"])
repetoire.append([6,"G5"])
repetoire.append([7,"C6"])
repetoire.append([8,"E6"])
repetoire.append([9,"G6"])
repetoire.append([10,"C7"])
repetoire.append([-1,"D4"])
repetoire.append([-2,"G4"])
repetoire.append([-3,"B4"])
repetoire.append([-4,"D5"])
repetoire.append([-5,"F5"])
repetoire.append([-6,"A5"])
repetoire.append([-7,"B5"])
repetoire.append([-8,"D6"])
repetoire.append([-9,"F6"])	
repetoire.append([-10,"A6"])

keys = []
keys.append(['K24', 'C1'])
keys.append(['K25', 'C#1'])
keys.append(['K26', 'D1'])
keys.append(['K27', 'D#1'])
keys.append(['K28', 'E1'])
keys.append(['K29', 'F1'])
keys.append(['K30', 'F#1'])
keys.append(['K31', 'G1'])
keys.append(['K32', 'G#1'])
keys.append(['K33', 'A1'])
keys.append(['K34', 'A#1'])
keys.append(['K35', 'B1'])
keys.append(['K36', 'C2'])
keys.append(['K37', 'C#2'])
keys.append(['K38', 'D2'])
keys.append(['K39', 'D#2'])
keys.append(['K40', 'E2'])
keys.append(['K41', 'F2'])
keys.append(['K42', 'F#2'])
keys.append(['K43', 'G2'])
keys.append(['K44', 'G#2'])
keys.append(['K45', 'A2'])
keys.append(['K46', 'A#2'])
keys.append(['K47', 'B2'])
keys.append(['K48', 'C3'])
keys.append(['K49', 'C#3'])
keys.append(['K50', 'D3'])
keys.append(['K51', 'D#3'])
keys.append(['K52', 'E3'])
keys.append(['K53', 'F3'])
keys.append(['K54', 'F#3'])
keys.append(['K55', 'G3'])
keys.append(['K56', 'G#3'])
keys.append(['K57', 'A3'])
keys.append(['K58', 'A#3'])
keys.append(['K59', 'B3'])
keys.append(['K60', 'C4'])
keys.append(['K61', 'C#4'])
keys.append(['K62', 'D4'])
keys.append(['K63', 'D#4'])
keys.append(['K64', 'E4'])
keys.append(['K65', 'F4'])
keys.append(['K66', 'F#4'])
keys.append(['K67', 'G4'])
keys.append(['K68', 'G#4'])
keys.append(['K69', 'A4'])
keys.append(['K70', 'A#4'])
keys.append(['K71', 'B4'])
keys.append(['K72', 'C5'])
keys.append(['K73', 'C#5'])
keys.append(['K74', 'D5'])
keys.append(['K75', 'D#5'])
keys.append(['K76', 'E5'])
keys.append(['K77', 'F5'])
keys.append(['K78', 'F#5'])
keys.append(['K79', 'G5'])
keys.append(['K80', 'G#5'])
keys.append(['K81', 'A5'])
keys.append(['K82', 'A#5'])
keys.append(['K83', 'B5'])
keys.append(['K84', 'C6'])
keys.append(['K85', 'C#6'])
keys.append(['K86', 'D6'])
keys.append(['K87', 'D#6'])
keys.append(['K88', 'E6'])
keys.append(['K89', 'F6'])
keys.append(['K90', 'F#6'])
keys.append(['K91', 'G6'])
keys.append(['K92', 'G#6'])
keys.append(['K93', 'A6'])
keys.append(['K94', 'A#6'])
keys.append(['K95', 'B6'])
keys.append(['K96', 'C7'])
keys.append(['K97', 'C#7'])
keys.append(['K98', 'D7'])
keys.append(['K99', 'D#7'])
keys.append(['K100', 'E7'])
keys.append(['K101', 'F7'])
keys.append(['K102', 'F#7'])
keys.append(['K103', 'G7'])
keys.append(['K104', 'G#7'])
keys.append(['K105', 'A7'])
keys.append(['K106', 'A#7'])
keys.append(['K107', 'B7'])
keys.append(['K108', 'C8'])
keys.append(['K109', 'C#8'])
keys.append(['K110', 'D8'])
keys.append(['K111', 'D#8'])
keys.append(['K112', 'E8'])
keys.append(['K113', 'F8'])



def get_track(inp):
	#track_to_get = get_track(inp)
	#track_to_get = get_track(title,song array)
	#track_to_get = get_track("prelude2",songs)
	title = inp[0]
	songs = inp[1]	
	track = ""
	for a,val in enumerate(songs):
		if title in val:
			track = val[0]
			break
	return track


def dataframe_to_list(inp):
	#array = dataframe_to_list([data as dataframe])
	#array = dataframe_to_list([data_dataframe])
	import pandas as pd
	data_dataframe = inp[0]
	if ".xls" in data_dataframe:
		data_dataframe = pd.read_excel(data_dataframe)
	#else:
	#	data_dataframe = load_file
	numpy_array = data_dataframe.to_numpy()
	normal_array = numpy_array.tolist()
	new_list = []
	for a,val in enumerate(normal_array):
		new_val = []
		for b,valb in enumerate(val):
			if type(valb)==float:
				if str(valb)=="nan":
					continue
			new_val.append(valb)
		new_list.append(new_val)
	normal_array = new_list	
	pri(normal_array)
	return normal_array

def load_data(inp):
	#data = load_data([load_file])
	load_file = inp[0]
	#loaded_data
	if ".xls" in load_file:
		import xlrd
		workbook = xlrd.open_workbook(load_file)
		sheet1 = workbook.sheet_by_index(0)	
		new_data = []
		for rowNumber in range(sheet1.nrows):
		    row = sheet1.row_values(rowNumber)
		    check = row[::-1]
		    new_row = []
		    for a,val in enumerate(check):
		    	if len(val)==0:
		    		continue
		    	new_row.append(val)
		    new_row = new_row[::-1]
		    new_data.append(new_row)
		    #pri(row)
		    #print(row)
		    #loaded_data.append(row)
		loaded_data = new_data
		#pri(loaded_data)
	if ".csv" in load_file:
		loaded_data = []
		with open(load_file, newline='') as csvfile:
			#make dat.csv in the ear folder, not in the cod folder
		    reader = csv.reader(csvfile)
		    for row in reader:
		    	loaded_data.append(row)
	if ".json" in load_file:
		with open(load_file, 'r') as f:
			loaded_data = json.load(f)
	if ".txt" in load_file or ".py" in load_file:
		#text = rea5(["aul2","py"])
		#che = "."+typ
		#if che not in nam:
		#	nam = nam+"."+typ
		text = open(load_file, "r")
		loaded_data = text.read()
		#return tex2
		#print(f.read())
	return loaded_data

def get_earnings_dates(inp):
	#max 25 a day
	#get previous earnings dates from alphavantage
	#get_earnings_dates([stock_list])
	stock_list = inp[0]
	import requests
	file_extension = ".json"	
	earn_list = os.listdir(earn_folder)
	key_alpha_vantange = "FK98MBU38O64HAMH"
	for a,val in enumerate(stock_list):
		symbol = val[2]
		check = symbol+"-previous_earnings_dates"+file_extension
		if check in earn_list:
			continue
		#print(check)
		#sym = val.replace(".json","")
		url = "https://www.alphavantage.co/query?function=EARNINGS&symbol="+symbol+ "&apikey="+key_alpha_vantange
		#print(symbol,a,val)
		alpha_information = requests.get(url)
		alpha_json = alpha_information.json()
		#sav = dire+"\\"+sym+ext
		save_file = earn_folder+"\\"+check
		with open(save_file, 'w') as output_file:
		    json.dump(alpha_json, output_file)
		print(symbol,"alpha vantage",save_file)
		sle([21])

def dictionary_to_list(inp):
	#new_list =dictionary_to_list([dictionary])
	#new_list =dictionary_to_list([val[4]])
	dictionary = inp[0]
	new_list = list(dictionary.values())
	return new_list
	#new_list = list(val[4].values())


def dictionary_to_list2(inp):
	#new_list =dictionary_to_list2([dictionary])
	#new_list =dictionary_to_list2([val[4]])
	dictionary = inp[0]
	for a,val in enumerate(dictionary):
		if a==0:
			header = []
		header.append(val)
	header = [header]
	new_list = list(dictionary.values())
	new_list = header+new_list
	for a,val in enumerate(new_list):
		print(type(val))
	return new_list

def dictionary_to_list3(inp):
	#covnert dictionary to list to 2nd degree?
	#new_list =dictionary_to_list3([dictionary])
	#new_list =dictionary_to_list3([val[4]])
	data_dictionary = inp[0]
	for a,val in enumerate(data_dictionary):
		if a==0:
			header = []
		header.append(val)
	header = [header]
	new_list = list(data_dictionary.values())
	new_list = header+new_list
	for a,val in enumerate(new_list):
		if type(val)==dict:
				new_list[a] = dictionary_to_list2([val])
		if type(val)==list:
			for b,valb in enumerate(val):
				if type(valb)==dict:
					new_list[a][b] = dictionary_to_list2([valb])
	pri(new_list)
	return new_list

def polygon_check(inp):
	#polygon_check([])
	file = "polygon.txt"
	text = load_data([file])
	times = nex4([text,"\n","\n"])
	for a,val in enumerate(times):
		times[a] = val.replace("\n","")
	#pri(times)
	from datetime import datetime
	current_time = datetime.now()
	import datetime
	one_min_ago = current_time+datetime.timedelta(seconds=-65)
	print("one min ago",one_min_ago)
	last_minute = 0
	print_last_five =0
	for a,val in enumerate(times):
		try:
			specific = datetime.datetime.strptime(val,"%H:%M:%S - %m/%d/%Y")
		except:
			continue
		if print_last_five<6:
			print("specific" ,specific)
		print_last_five = print_last_five+1
		if specific>one_min_ago:
			last_minute = last_minute+1
			if last_minute==5:
				print("too many requests..must wait..")
	print("last minute = ",last_minute)
	print(str(last_minute)+" requests last minute")
	append_time = current_time.strftime("%H:%M:%S - %m/%d/%Y")
	#current_time = datetime.now()
	text = append_time+"\n"+text
	#print(text)
	#print(current_time)
	wri2([file,text])	

def trading_inputs(inp):	
	#trading_data = trading_inputs([])
	loaded_text = load_data(["trading.txt"])
	print(loaded_text)
	lines = nex4([loaded_text,"\n","\n"])
	inputs = []
	new_list = []
	for a,val in enumerate(lines):
		val=val.replace("\n","")
		if len(val)==0:
			continue
		val2 = val.find("=")
		val = val[val2:len(val)]
		val = val.replace("=","")
		val = val.replace(" ","")
		val = val.upper()
		new_list.append(val)
	pri(new_list)
	return new_list


#exec(open('util.py').read())
from tkinter import *
import os
def deriv2(inp):
	import requests
	import numpy as np
	from datetime import datetime
	trading_data = inp
	symbol = trading_data[0]
	date_in_question = trading_data[1]
	call_or_put = trading_data[2]
	override_expiration = trading_data[3]
	"""
	#old input method
	# if len override_price>0, then price overrides
	trading_data = trading_inputs([])
	symbol = trading_data[0]
	date_in_question = trading_data[1]
	call_or_put = trading_data[2]
	override_expiration = trading_data[3]
	#date_in_question = "2024-03-08"
	#call_or_put = "c"
	#override_expiration = "515"
	"""
	#date_in_question = "2023-11-29"
	#dat2 = datetime.datetime.strptime(dat,'%b %d, %Y')
	from datetime import datetime
	date_object = datetime.strptime(date_in_question,'%Y-%m-%d')
	day_of_week = date_object.weekday()
	if day_of_week==4:
		#next_friday = date_in_question
		next_friday = datetime.strptime(date_in_question,'%Y-%m-%d')
	if day_of_week!=4:
		import datetime
		if day_of_week<4:
			next_friday = date_object+datetime.timedelta(4-day_of_week)
		if day_of_week>4:
			next_friday =date_object+datetime.timedelta(11-day_of_week) 
	print(date_in_question)
	print(day_of_week)
	print(next_friday)
	if len(symbol)==0:
		symbol =str(input("symbol = "))
	move_week = 0
	#if len(date_of_options)==0:
	#	move_week = int(str(input("move week = ")))
	#override_price = "20"
	if len(override_expiration)==0:
		override_expiration =str(input("override_price = "))
	symbol = symbol.upper()
	call_or_put = call_or_put.upper()
	price = override_expiration
	#symbol = "AMD"
	#override_price = "20"
	#symbol = "SPY"
	#start_date = "2024-01-26"
	#start_date = "2024-02-01"
	#start_date = "2024-02-08"
	#start_date = "2024-02-16"
	#start_date = "2024-02-23"
	#start_date = "2024-03-01"
	#start_date = "2024-03-08"
	start_date = date_in_question
	from datetime import datetime
	start_date_object = datetime.strptime(start_date,'%Y-%m-%d')
	#start_date_object = start_date
	import datetime
	start_date_delta = start_date_object-datetime.timedelta(7*move_week)
	start_date_converted = start_date_delta.strftime('%Y-%m-%d')
	print(start_date_converted)
	start_date = start_date_converted
	#sye()
	from datetime import datetime
	#sye()
	end_date = start_date_converted
	#end_date = "2024-02-23"
	folder = os.getcwd()
	earn_folder = folder+"\\earn\\"
	earn_list = os.listdir(earn_folder)
	file_name = symbol+"."+start_date+".json"
	save_file = earn_folder+file_name
	#print(file_name)
	#sye()
	if file_name not in earn_list:
		#get stock prices
		url1 = "https://api.polygon.io/v2/aggs/ticker/"
		url2 = "/range/1/minute/"
		url3 = "/"
		url4 = "?adjusted=true&sort=asc&limit=500&apiKey=65JaxrhDSYET1StvPxZy1KgpnttWna98"
		url = url1+symbol+url2+start_date+url3+end_date+url4
		#getting data from polygon with json
		polygon_check([])
		data_back = requests.get(url)
		json_data = data_back.json()
		#print("-----------------have done request to polygon--------------------")
		with open(save_file, 'w') as f:
		    json.dump(json_data, f)
		wri2([save_file,json_data])
		print("output file = ",save_file)
		print("have gotten stock data..")
		sle([5])
	price_info = load_data([save_file])
	#print("save file",save_file)
	minute_data = price_info["results"]
	if len(str(price))==0:
		price = minute_data[4]["c"]
	print ("price = ",price)
	#if len(str(price))==0:
	#	price = float(override_price)	
	price2 = round(float(price))
	options_code = str(price2)+"000"
	for a in range(0,8-len(options_code)):
		options_code = "0"+options_code
	options_code =call_or_put+options_code
	#file_name = symbol+"."+start_date+"."+options_code+".json"
	#save_file = earn_folder+file_name
	print(file_name)
	url_options1 = "https://api.polygon.io/v2/aggs/ticker/O:"
	#start_date_object = datetime.strptime(start_date,'%Y-%m-%d')
	expiration_year = next_friday.strftime("%y")
	expiration_month = next_friday.strftime("%m")
	expiration_day = next_friday.strftime("%d")
	options_date = expiration_year+expiration_month+expiration_day
	options_code = options_date+options_code
	url_options2 = "/range/1/minute/"
	url_options3 = "/"
	url_options4 = "?adjusted=true&sort=asc&limit=500&apiKey=65JaxrhDSYET1StvPxZy1KgpnttWna98"
	url_options = url_options1+symbol+options_code+url_options2+start_date+url_options3+end_date+url_options4
	#getting data from polygon with json
	print(url_options)
	print(options_code)
	file_name = symbol+"."+start_date+"."+options_code+".json"
	save_file = earn_folder+file_name
	print(file_name)
	print(save_file)
	#sye()
	if file_name not in earn_list:
		polygon_check([])
		data_back = requests.get(url_options)
		json_data = data_back.json()
		#print("-----------------have done request to polygon--------------------")
		with open(save_file, 'w') as f:
		    json.dump(json_data, f)
		print("output file = ",save_file)
		print("now gotten options data..")
	data_dictionary = load_data([save_file])
	#pri(data_dictionary['results'])
	options_data = []
	skip_early = 0
	for a,val in enumerate(data_dictionary['results']):
		more = []
		timestamp_integer = int(val["t"])
		timestamp_integer = (timestamp_integer)/1000
		timestamp_object = datetime.fromtimestamp(timestamp_integer)
		converted_time = timestamp_object.strftime("%m/%d/%Y-%H:%M")
		if "09:30" in converted_time:
			skip_early = 1
		#if skip_early==0:
		#	continue
		more.append(converted_time)
		more.append(val["o"])
		more.append(val["h"])
		more.append(val["l"])
		more.append(val["c"])
		options_data.append(more)
		if "16:00" in converted_time:
			break
	pri(options_data)
	import pandas as pd
	dataframe = pd.DataFrame(options_data)
	#print(dataframe)
	import plotly.graph_objects as go
	from datetime import datetime
	#df = pd.read_csv('https://raw.githubusercontent.com/plotly/datasets/master/finance-charts-apple.csv')
	print(file_name)
	print(dataframe)
	#sye()
	fig = go.Figure(data=[go.Candlestick(x=dataframe[0],
					open=dataframe[1],
	                high=dataframe[2],
	                low=dataframe[3],
	                close=dataframe[4])])
	fig.update_layout(
    title=file_name
    )
	fig.show()

#interface used to run it..
import tkinter
import tkinter as tk
master = Tk()
master.title("GUI")
master.geometry("200x120")
Label(master, text='Symbol').grid(row=0)
Label(master, text='Date in Question').grid(row=1)
Label(master, text='Call or Put').grid(row=2)
Label(master, text='Override Expiration').grid(row=3)
Label(master, text='Get info').grid(row=4)
e1 = Entry(master)
e1.insert(END,"SPY") 
e2 = Entry(master)
e2.insert(END,"2024-03-07")
e3 = Entry(master)
e3.insert(END,"p")
e4 = Entry(master)
e4.insert(END,"510")
e1.grid(row=0, column=1)
e2.grid(row=1, column=1)
e3.grid(row=2, column=1)
e4.grid(row=3, column=1)
def run_whole_thing():
	#symbol = e1
	#print(symbol)
	symbol = e1.get()
	date_in_question = e2.get()
	call_or_put = e3.get()
	override_expiration = e4.get()
	inputs = [symbol,date_in_question,call_or_put,override_expiration]
	deriv2(inputs)
B = Button(master, text ="Get info", command = run_whole_thing)
B.place(x=110,y=85)
print(e1,e2)
mainloop()